
## This is the THIRD STEP in the craniometrics database workflow that will generate a population file
## Note that unlike the previous steps, there is a line in this code at the very end that you'll want to update!

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import coordinate_to_tuple

# Directory containing the .xlsm files
directory = '***'

# Specify the ranges and their corresponding sheets to read
specific_ranges = {
    'Height': ['G2:I82', 'L2:L82'],
    'Anterior': ['G2:J91'],
    'Posterior': ['G2:J91'],
    'AntPost': ['G2:I181', 'L2:L181'],
    'Right': ['G2:J91'],
    'Left': ['G2:J91'],
    'LR': ['G2:I181', 'L2:L181']
}

# Initialize a dictionary to store the sums and counts for averaging
data = {sheet: {} for sheet in specific_ranges}

# Read each .xlsm file in the directory
for filename in os.listdir(directory):
    if filename.endswith('.xlsm') and not filename.startswith('~$'):
        filepath = os.path.join(directory, filename)
        xls = pd.ExcelFile(filepath)
        print(f"Reading file: {filename}")
        
        # Read specified ranges from each sheet
        for sheet, ranges in specific_ranges.items():
            if sheet in xls.sheet_names:
                print(f"Processing sheet: {sheet} in file: {filename}")
                df = xls.parse(sheet)
                print(f"DataFrame loaded for {sheet} with shape: {df.shape}")
                
                for range_ in ranges:
                    print(f"Processing range: {range_} in sheet: {sheet}")
                    start_cell, end_cell = range_.split(':')
                    start_row, start_col = coordinate_to_tuple(start_cell)
                    end_row, end_col = coordinate_to_tuple(end_cell)
                    start_row -= 1  # adjust for 0-based index
                    start_col -= 1  # adjust for 0-based index
                    end_row = min(end_row - 1, df.shape[0] - 1)  # adjust for 0-based index and bounds
                    end_col = min(end_col - 1, df.shape[1] - 1)  # adjust for 0-based index and bounds
                    
                    for row in range(start_row, end_row + 1):
                        for col in range(start_col, end_col + 1):
                            value = df.iat[row, col]
                            if pd.notna(value) and isinstance(value, (int, float)):
                                print(f"Read numeric value: {value} at position ({row+1},{col+1})")
                                if value != 0:  # Ignore zero values
                                    cell_key = (row, col)
                                    if cell_key not in data[sheet]:
                                        data[sheet][cell_key] = {'sum': 0, 'count': 0}
                                    data[sheet][cell_key]['sum'] += value
                                    data[sheet][cell_key]['count'] += 1
                            else:
                                print(f"Non-numeric or NaN value ignored at position ({row+1},{col+1}): {value}")

# Debugging output: Check the contents of the data dictionary
print("Data dictionary contents after processing all files:")
for sheet, sheet_data in data.items():
    print(f"Sheet: {sheet}, Data points: {len(sheet_data)}")
    for cell_key, value_dict in sheet_data.items():
        print(f"Cell: {cell_key}, Sum: {value_dict['sum']}, Count: {value_dict['count']}")

# Create a new workbook and write the averaged data to it
output_wb = Workbook()
for sheet_name in specific_ranges.keys():
    ws = output_wb.create_sheet(title=sheet_name)
    print(f"Writing data to sheet: {sheet_name}")
    for (row, col), value_dict in data[sheet_name].items():
        if value_dict['count'] > 0:
            avg_value = value_dict['sum'] / value_dict['count']
            ws.cell(row=row + 1, column=col + 1, value=avg_value)
            print(f"Written avg value: {avg_value} at position ({row+1},{col+1}) in sheet {sheet_name}")
        else:
            print(f"No data to write at position ({row+1},{col+1}) in sheet {sheet_name}")

# Remove the default sheet created by Workbook
if 'Sheet' in output_wb.sheetnames:
    del output_wb['Sheet']

# Save the new workbook with the averaged values
output_filepath = '***'
output_wb.save(output_filepath)

print(f"Averaged data has been written to {output_filepath}")
