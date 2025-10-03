## This is the second part to the flipping procedure for the craniometrics data processing workflow
## For this script, you want to ensure you udpate the directories appropriately to only flip the data that is left-sided

import pandas as pd
import openpyxl
import os
import shutil
from openpyxl.utils.dataframe import dataframe_to_rows

# Define source and destination directories
source_dir = r"C:\Users\akane\Desktop\cm_data\UC_points_alignedR"
destination_dir = r"C:\Users\akane\Desktop\cm_data\UC_verifiedR"

# Create destination directory if it doesn't exist
os.makedirs(destination_dir, exist_ok=True)

# Function to read a sheet as values
def read_sheet_as_values(workbook, sheet_name):
    sheet = workbook[sheet_name]
    data = sheet.values
    cols = next(data)[0:]
    data = list(data)
    df = pd.DataFrame(data, columns=cols)
    return df

# Function to write a DataFrame to a sheet, replacing all formulas with values
def write_values_to_sheet(workbook, df, sheet_name):
    sheet = workbook[sheet_name]
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx, column=c_idx, value=value)

# Function to swap data between sheets "Left" and "Right" and update "LR"
def swap_sheets_and_save(file_path, destination_dir):
    try:
        print(f"Processing {file_path}")
        
        # Copy the original file to the destination directory
        new_file_path = os.path.join(destination_dir, os.path.basename(file_path))
        shutil.copyfile(file_path, new_file_path)
        
        # Load the copied workbook with keep_vba=True to preserve macros
        workbook = openpyxl.load_workbook(new_file_path, keep_vba=True)
        
        # Load data from sheets
        left_df = read_sheet_as_values(workbook, 'Left')
        right_df = read_sheet_as_values(workbook, 'Right')
        lr_df = read_sheet_as_values(workbook, 'LR')

        # Swap data between "Left" and "Right"
        temp_df = left_df.copy()
        left_df = right_df.copy()
        right_df = temp_df.copy()
        
        # Update "LR" sheet
        for idx, row in lr_df.iterrows():
            if row['Intersection'] == 'Right':
                lr_df.loc[idx, 'Intersection'] = 'Left'
            elif row['Intersection'] == 'Left':
                lr_df.loc[idx, 'Intersection'] = 'Right'
        
        # Clear the original data in the sheets
        left_sheet = workbook['Left']
        right_sheet = workbook['Right']
        lr_sheet = workbook['LR']
        
        left_sheet.delete_rows(2, left_sheet.max_row)
        right_sheet.delete_rows(2, right_sheet.max_row)
        lr_sheet.delete_rows(2, lr_sheet.max_row)
        
        # Write the new data to the sheets
        write_values_to_sheet(workbook, left_df, 'Left')
        write_values_to_sheet(workbook, right_df, 'Right')
        write_values_to_sheet(workbook, lr_df, 'LR')
        
        # Save the modified workbook
        workbook.save(new_file_path)
        
        print(f"Swapped data and saved to {new_file_path}")
    except Exception as e:
        print(f"Error processing {file_path}: {e}")

# Iterate through the files in the source directory
for file_name in os.listdir(source_dir):
    if file_name.endswith('.xlsm'):
        file_path = os.path.join(source_dir, file_name)
        swap_sheets_and_save(file_path, destination_dir)
